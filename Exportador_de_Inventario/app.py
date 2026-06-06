#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Exportador Inteligente de Inventarios
=====================================
Sistema de exportación y transformación de archivos Excel que permite cargar
documentos de inventario con estructuras diferentes y reorganizarlos automáticamente
según la estructura definida en el archivo modelo.

Tecnologías:
- Flask (Backend)
- Pandas (Manipulación de datos)
- OpenPyXL (Lectura/Escritura Excel)
- HTML5/CSS3/JavaScript (Frontend)

Autor: Sistema Automático
Fecha: 2026-06-02
"""

import os
import re
import json
import uuid
from datetime import datetime
from difflib import SequenceMatcher
from collections import defaultdict

from flask import Flask, render_template, request, jsonify, send_file, session
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURACIÓN DE LA APLICACIÓN
# =============================================================================
app = Flask(__name__)
app.secret_key = 'exportador-inventarios-secret-key-2026'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max

# Directorios
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
OUTPUT_FOLDER = os.path.join(BASE_DIR, 'outputs')
MODEL_FILE = os.path.join(BASE_DIR, 'modelo', 'Estructura de distriapp.xlsx')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(os.path.dirname(MODEL_FILE), exist_ok=True)

# =============================================================================
# ESTRUCTURA OBJETIVO (Columnas del archivo modelo)
# =============================================================================
TARGET_COLUMNS = [
    'sku', 'locale', 'attribute_family_code', 'type', 'categories',
    'name', 'description', 'short_description', 'status', 'visible_individually',
    'brand', 'guest_checkout', 'price', 'Iva', 'url_key'
]

# Tipos de datos esperados por columna
TARGET_DTYPES = {
    'sku': str,
    'locale': str,
    'attribute_family_code': str,
    'type': str,
    'categories': str,
    'name': str,
    'description': str,
    'short_description': str,
    'status': int,
    'visible_individually': int,
    'brand': str,
    'guest_checkout': int,
    'price': int,
    'Iva': int,
    'url_key': str
}

# Valores por defecto para columnas faltantes
DEFAULT_VALUES = {
    'sku': '',
    'locale': 'es',
    'attribute_family_code': 'default',
    'type': 'simple',
    'categories': 'General',
    'name': '',
    'description': '',
    'short_description': '',
    'status': 1,
    'visible_individually': 1,
    'brand': '',
    'guest_checkout': 0,
    'price': 0,
    'Iva': 19,
    'url_key': ''
}

# =============================================================================
# DICCIONARIO DE MAPEO INTELIGENTE
# =============================================================================
# Mapea posibles nombres de columnas de entrada a las columnas objetivo
COLUMN_ALIASES = {
    'sku': [
        'sku', 'codigo', 'código', 'codigo producto', 'código producto', 'id', 'id producto',
        'referencia', 'ref', 'producto', 'item', 'item code', 'product code', 'cod',
        'codigo sku', 'código sku', 'sku code', 'clave', 'clave producto', 'número',
        'numero', 'numero producto', 'número producto', 'id_articulo', 'id articulo',
        'articulo', 'artículo', 'product_id', 'product id', 'item_id', 'item id'
    ],
    'locale': [
        'locale', 'idioma', 'language', 'lang', 'loc', 'regional', 'region', 'pais', 'país',
        'country', 'idioma producto', 'lenguaje', 'localizacion', 'localización'
    ],
    'attribute_family_code': [
        'attribute_family_code', 'familia', 'family', 'attribute family', 'atributo familia',
        'familia atributo', 'attribute_family', 'attr_family', 'attr family', 'tipo familia',
        'categoria familia', 'categoría familia', 'grupo', 'group', 'attribute group'
    ],
    'type': [
        'type', 'tipo', 'product type', 'tipo producto', 'tipo articulo', 'tipo artículo',
        'categoria', 'categoría', 'category', 'class', 'clase', 'clase producto',
        'tipo_item', 'tipo item', 'item_type', 'item type'
    ],
    'categories': [
        'categories', 'categorias', 'categorías', 'category', 'categoria', 'categoría',
        'grupo', 'group', 'grupo producto', 'linea', 'línea', 'linea producto', 'línea producto',
        'departamento', 'depto', 'seccion', 'sección', 'familia', 'subcategoria', 'subcategoría',
        'subcategory', 'clasificacion', 'clasificación'
    ],
    'name': [
        'name', 'nombre', 'nombre producto', 'producto', 'articulo', 'artículo',
        'descripcion corta', 'descripción corta', 'titulo', 'título', 'title',
        'product name', 'nombre articulo', 'nombre artículo', 'denominacion', 'denominación',
        'desc producto', 'desc articulo', 'desc artículo', 'item name', 'item_name',
        'producto nombre', 'articulo nombre', 'artículo nombre'
    ],
    'description': [
        'description', 'descripcion', 'descripción', 'desc', 'descripcion larga',
        'descripción larga', 'detalle', 'detalle producto', 'especificacion', 'especificación',
        'especificaciones', 'caracteristicas', 'características', 'info', 'informacion',
        'información', 'notas', 'observaciones', 'detalles', 'desc completa',
        'descripcion completa', 'descripción completa', 'full description'
    ],
    'short_description': [
        'short_description', 'descripcion corta', 'descripción corta', 'resumen',
        'breve descripcion', 'breve descripción', 'desc corta', 'short desc',
        'mini descripcion', 'mini descripción', 'extracto', 'sumario', 'preview',
        'resumen producto', 'desc breve', 'descripcion breve', 'descripción breve'
    ],
    'status': [
        'status', 'estado', 'activo', 'active', 'habilitado', 'enabled', 'disponible',
        'available', 'publicado', 'published', 'visible', 'estado producto', 'situacion',
        'situación', 'condicion', 'condición', 'state', 'item status', 'product status',
        'estado articulo', 'estado artículo', 'activo producto', 'producto activo'
    ],
    'visible_individually': [
        'visible_individually', 'visible individual', 'visible_individual', 'individual',
        'visibilidad', 'visibility', 'mostrar', 'show', 'display', 'mostrar individual',
        'visible solo', 'solo visible', 'individualmente', 'individualmente visible',
        'visibilidad producto', 'mostrar producto', 'producto visible', 'item visible'
    ],
    'brand': [
        'brand', 'marca', 'fabricante', 'manufacturer', 'maker', 'marca producto',
        'product brand', 'marca articulo', 'marca artículo', 'nombre marca', 'brand name',
        'linea marca', 'línea marca', 'proveedor marca', 'marca proveedor'
    ],
    'guest_checkout': [
        'guest_checkout', 'guest checkout', 'invitado', 'guest', 'checkout', 'compra invitado',
        'permitir invitado', 'allow guest', 'guest purchase', 'compra sin registro',
        'sin registro', 'no registro', 'anonymous', 'anonimo', 'anónimo', 'guest_allowed',
        'guest allowed', 'permitir guest'
    ],
    'price': [
        'price', 'precio', 'precio unitario', 'valor', 'valor unitario', 'costo', 'costo unitario',
        'precio venta', 'precio de venta', 'venta', 'pvp', 'p.v.p', 'p.v.p.',
        'precio producto', 'precio articulo', 'precio artículo', 'item price', 'unit price',
        'precio_unidad', 'precio unidad', 'valor producto', 'amount', 'monto', 'importe',
        'precio lista', 'precio de lista', 'tarifa', 'rate', 'tarifa unitaria',
        'precio sin iva', 'precio sin impuesto', 'neto', 'precio neto', 'subtotal'
    ],
    'Iva': [
        'iva', 'impuesto', 'tax', 'vat', 'impuesto valor agregado', 'valor agregado',
        'tasa impuesto', 'tasa iva', 'porcentaje iva', '% iva', 'impuesto producto',
        'tax rate', 'vat rate', 'rate iva', 'iva rate', 'impuesto ventas', 'impuesto a las ventas',
        'tax amount', 'iva amount', 'monto iva', 'valor iva'
    ],
    'url_key': [
        'url_key', 'url key', 'url', 'slug', 'permalink', 'enlace', 'link', 'url producto',
        'product url', 'url articulo', 'url artículo', 'direccion', 'dirección', 'path',
        'ruta', 'route', 'url path', 'product link', 'item url', 'item_url', 'product_url',
        'url amigable', 'friendly url', 'seo url', 'url seo'
    ]
}

# =============================================================================
# FUNCIONES DE UTILIDAD
# =============================================================================

def normalize_text(text):
    """Normaliza texto para comparación: minúsculas, sin acentos, sin espacios extra."""
    if pd.isna(text) or text is None:
        return ''
    text = str(text).strip().lower()
    # Reemplazar acentos
    replacements = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'à': 'a', 'è': 'e', 'ì': 'i', 'ò': 'o', 'ù': 'u',
        'ä': 'a', 'ë': 'e', 'ï': 'i', 'ö': 'o', 'ü': 'u',
        'ñ': 'n', 'ç': 'c', 'º': '', 'ª': ''
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    # Eliminar caracteres especiales y espacios múltiples
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def similarity_score(a, b):
    """Calcula la similitud entre dos textos (0-1)."""
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, normalize_text(a), normalize_text(b)).ratio()


def find_best_match(column_name, target_aliases, threshold=0.6):
    """Encuentra la mejor coincidencia entre un nombre de columna y los alias conocidos."""
    normalized_col = normalize_text(column_name)
    best_target = None
    best_score = 0.0

    for target, aliases in target_aliases.items():
        for alias in aliases:
            # Coincidencia exacta normalizada
            if normalized_col == normalize_text(alias):
                return target, 1.0
            # Coincidencia parcial
            score = similarity_score(column_name, alias)
            if score > best_score and score >= threshold:
                best_score = score
                best_target = target

    return best_target, best_score


def clean_numeric_value(value, default=0):
    """Limpia y convierte valores numéricos."""
    if pd.isna(value) or value is None:
        return default
    if isinstance(value, (int, float)):
        return int(value) if value == int(value) else float(value)
    # Limpiar string
    cleaned = str(value).strip()
    # Eliminar símbolos de moneda y separadores
    cleaned = re.sub(r'[$€£¢]', '', cleaned)
    cleaned = cleaned.replace(',', '')
    try:
        return int(float(cleaned))
    except (ValueError, TypeError):
        return default


def clean_text_value(value, default=''):
    """Limpia valores de texto."""
    if pd.isna(value) or value is None:
        return default
    text = str(value).strip()
    if text.lower() in ['nan', 'null', 'none', '']:
        return default
    return text


def infer_boolean_value(value):
    """Infiere valor booleano numérico (0/1) desde diversos formatos."""
    if pd.isna(value) or value is None:
        return 0
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return 1 if value != 0 else 0
    text = str(value).strip().lower()
    positive = ['si', 'sí', 'yes', 'true', '1', 'activo', 'active', 'enabled', 
                'habilitado', 'disponible', 'available', 'publicado', 'published',
                'visible', 'mostrar', 'show', 'verdadero', 'v']
    negative = ['no', 'false', '0', 'inactivo', 'inactive', 'disabled', 
                'deshabilitado', 'no disponible', 'unavailable', 'no publicado',
                'unpublished', 'oculto', 'hidden', 'falso', 'f']
    if text in positive:
        return 1
    if text in negative:
        return 0
    return 1 if text else 0


def generate_url_key(name, sku=''):
    """Genera una URL key a partir del nombre o SKU."""
    base = str(name) if name and not pd.isna(name) else str(sku)
    url = normalize_text(base)
    url = re.sub(r'\s+', '-', url)
    url = re.sub(r'-+', '-', url)
    url = url.strip('-')
    if not url and sku:
        url = str(sku).lower().strip()
    return url[:100]  # Limitar longitud


# =============================================================================
# FUNCIONES PRINCIPALES DE PROCESAMIENTO
# =============================================================================

def analyze_file_structure(file_path):
    """Analiza la estructura de un archivo Excel de entrada."""
    try:
        # Intentar leer con pandas primero
        df = pd.read_excel(file_path, nrows=5)

        # Si no hay columnas o están vacías, intentar sin encabezado
        if df.columns.str.contains('Unnamed').all() or len(df.columns) == 0:
            df = pd.read_excel(file_path, header=None, nrows=5)
            # Intentar detectar fila de encabezado
            for i in range(min(10, len(df))):
                row = df.iloc[i]
                if row.notna().sum() > len(row) * 0.5:
                    df = pd.read_excel(file_path, header=i)
                    break

        columns_info = []
        for col in df.columns:
            sample_values = df[col].dropna().head(3).tolist()
            columns_info.append({
                'original_name': str(col),
                'normalized_name': normalize_text(str(col)),
                'sample_values': [str(v)[:50] for v in sample_values],
                'data_type': str(df[col].dtype)
            })

        return {
            'success': True,
            'columns': columns_info,
            'total_columns': len(df.columns),
            'sample_rows': len(df)
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }


def auto_map_columns(input_columns, target_aliases, threshold=0.6):
    """Mapea automáticamente columnas de entrada a columnas objetivo."""
    mapping = {}  # target -> input_column
    used_inputs = set()
    confidence_scores = {}

    # Primera pasada: coincidencias exactas y de alta confianza
    for target in TARGET_COLUMNS:
        best_input = None
        best_score = 0.0

        for input_col in input_columns:
            if input_col in used_inputs:
                continue

            score = similarity_score(input_col, target)

            # Verificar aliases
            if target in target_aliases:
                for alias in target_aliases[target]:
                    alias_score = similarity_score(input_col, alias)
                    score = max(score, alias_score)

            if score > best_score and score >= threshold:
                best_score = score
                best_input = input_col

        if best_input:
            mapping[target] = best_input
            used_inputs.add(best_input)
            confidence_scores[target] = round(best_score, 2)

    # Segunda pasada: inferencia por contenido de muestra
    for target in TARGET_COLUMNS:
        if target in mapping:
            continue

        # Buscar columnas no usadas que puedan contener el tipo de dato esperado
        for input_col in input_columns:
            if input_col in used_inputs:
                continue

            # Heurísticas específicas por tipo de dato
            normalized = normalize_text(input_col)

            if target == 'price' and any(kw in normalized for kw in ['precio', 'valor', 'costo', 'price', 'amount']):
                mapping[target] = input_col
                used_inputs.add(input_col)
                confidence_scores[target] = 0.5
                break
            elif target == 'sku' and any(kw in normalized for kw in ['cod', 'sku', 'ref', 'id', 'item']):
                mapping[target] = input_col
                used_inputs.add(input_col)
                confidence_scores[target] = 0.5
                break
            elif target == 'name' and any(kw in normalized for kw in ['nombre', 'name', 'producto', 'articulo', 'desc']):
                mapping[target] = input_col
                used_inputs.add(input_col)
                confidence_scores[target] = 0.5
                break
            elif target == 'brand' and any(kw in normalized for kw in ['marca', 'brand', 'fabricante']):
                mapping[target] = input_col
                used_inputs.add(input_col)
                confidence_scores[target] = 0.5
                break

    return mapping, confidence_scores


def transform_data(df_input, column_mapping, confidence_scores):
    """Transforma los datos de entrada al formato objetivo."""
    df_output = pd.DataFrame()
    warnings = []

    for target in TARGET_COLUMNS:
        if target in column_mapping:
            input_col = column_mapping[target]

            if target in ['status', 'visible_individually', 'guest_checkout']:
                # Campos booleanos numéricos
                df_output[target] = df_input[input_col].apply(infer_boolean_value)
            elif target in ['price', 'Iva']:
                # Campos numéricos
                df_output[target] = df_input[input_col].apply(lambda x: clean_numeric_value(x, DEFAULT_VALUES[target]))
            elif target == 'url_key':
                # Generar URL key si no hay mapeo directo o está vacío
                if input_col in df_input.columns:
                    df_output[target] = df_input.apply(
                        lambda row: generate_url_key(row[input_col], row.get(column_mapping.get('sku', ''), '')),
                        axis=1
                    )
                else:
                    df_output[target] = df_input.apply(
                        lambda row: generate_url_key(row.get(column_mapping.get('name', ''), ''), 
                                                      row.get(column_mapping.get('sku', ''), '')),
                        axis=1
                    )
            else:
                # Campos de texto
                df_output[target] = df_input[input_col].apply(lambda x: clean_text_value(x, DEFAULT_VALUES[target]))

            if confidence_scores.get(target, 1.0) < 0.8:
                warnings.append({
                    'column': target,
                    'mapped_from': input_col,
                    'confidence': confidence_scores[target],
                    'message': f'Columna "{target}" mapeada desde "{input_col}" con confianza baja ({confidence_scores[target]})'
                })
        else:
            # Columna no mapeada - usar valor por defecto
            df_output[target] = DEFAULT_VALUES[target]
            warnings.append({
                'column': target,
                'mapped_from': None,
                'confidence': 0,
                'message': f'Columna "{target}" no encontrada. Se usará valor por defecto: {DEFAULT_VALUES[target]}'
            })

    return df_output, warnings


def style_excel_output(output_path, df):
    """Aplica estilos profesionales al archivo Excel de salida."""
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    ws.title = "Inventario Estandarizado"

    # Estilos
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    cell_font = Font(name='Calibri', size=10)
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    border = Border(
        left=Side(style='thin', color='BDBDBD'),
        right=Side(style='thin', color='BDBDBD'),
        top=Side(style='thin', color='BDBDBD'),
        bottom=Side(style='thin', color='BDBDBD')
    )

    # Aplicar estilos a encabezados
    for col_num, column in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

        # Ajustar ancho de columna
        max_length = len(str(column))
        for row in ws.iter_rows(min_row=2, max_row=min(50, ws.max_row), min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 50))
        ws.column_dimensions[get_column_letter(col_num)].width = max(max_length + 2, 12)

    # Aplicar estilos a celdas de datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border

    # Congelar paneles
    ws.freeze_panes = 'A2'

    # Filtros
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    return output_path


# =============================================================================
# RUTAS DE LA APLICACIÓN FLASK
# =============================================================================

@app.route('/')
def index():
    """Página principal - Interfaz de carga."""
    return render_template('index.html')


@app.route('/api/analyze', methods=['POST'])
def api_analyze():
    """API: Analiza un archivo Excel subido y devuelve su estructura."""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No se envió ningún archivo'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'Nombre de archivo vacío'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Formato no soportado. Use .xlsx o .xls'}), 400

    # Guardar archivo temporalmente
    file_id = str(uuid.uuid4())[:8]
    filename = f"{file_id}_{file.filename}"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    # Analizar estructura
    result = analyze_file_structure(filepath)

    if result['success']:
        # Generar mapeo automático
        input_columns = [col['original_name'] for col in result['columns']]
        mapping, confidence = auto_map_columns(input_columns, COLUMN_ALIASES)

        result['suggested_mapping'] = mapping
        result['confidence_scores'] = confidence
        result['file_id'] = file_id
        result['filename'] = file.filename

        # Guardar en sesión
        session['current_file'] = filepath
        session['current_file_id'] = file_id
        session['current_filename'] = file.filename

    return jsonify(result)


@app.route('/api/process', methods=['POST'])
def api_process():
    """API: Procesa el archivo actual y genera el Excel estandarizado."""
    data = request.get_json() or {}
    custom_mapping = data.get('custom_mapping', {})

    filepath = session.get('current_file')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'No hay archivo cargado. Sube un archivo primero.'}), 400

    try:
        # Leer archivo
        df_input = pd.read_excel(filepath)

        # Si no hay columnas válidas, intentar detectar encabezado
        if df_input.columns.str.contains('Unnamed').all():
            df_input = pd.read_excel(filepath, header=None)
            for i in range(min(10, len(df_input))):
                if df_input.iloc[i].notna().sum() > len(df_input.columns) * 0.3:
                    df_input = pd.read_excel(filepath, header=i)
                    break

        # Usar mapeo personalizado si se proporciona, sino automático
        if custom_mapping:
            column_mapping = {k: v for k, v in custom_mapping.items() if v}
        else:
            input_columns = list(df_input.columns)
            column_mapping, confidence_scores = auto_map_columns(input_columns, COLUMN_ALIASES)

        # Transformar datos
        df_output, warnings = transform_data(df_input, column_mapping, {})

        # Generar archivo de salida
        output_id = str(uuid.uuid4())[:8]
        output_filename = f"inventario_estandarizado_{output_id}.xlsx"
        output_path = os.path.join(OUTPUT_FOLDER, output_filename)

        # Guardar con pandas primero
        df_output.to_excel(output_path, index=False, sheet_name='Inventario')

        # Aplicar estilos con openpyxl
        style_excel_output(output_path, df_output)

        # Guardar en sesión
        session['output_file'] = output_path
        session['output_filename'] = output_filename

        # Preparar preview
        preview_data = df_output.head(5).to_dict('records')

        return jsonify({
            'success': True,
            'output_id': output_id,
            'output_filename': output_filename,
            'total_rows': len(df_output),
            'total_columns': len(df_output.columns),
            'column_mapping': column_mapping,
            'warnings': warnings,
            'preview': preview_data
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@app.route('/api/download/<output_id>')
def api_download(output_id):
    """Descarga el archivo procesado."""
    output_path = session.get('output_file')
    output_filename = session.get('output_filename', 'inventario_estandarizado.xlsx')

    if not output_path or not os.path.exists(output_path):
        # Buscar en carpeta de outputs
        for f in os.listdir(OUTPUT_FOLDER):
            if f.startswith(f"inventario_estandarizado_{output_id}"):
                output_path = os.path.join(OUTPUT_FOLDER, f)
                output_filename = f
                break

    if not output_path or not os.path.exists(output_path):
        return jsonify({'success': False, 'error': 'Archivo no encontrado'}), 404

    return send_file(
        output_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=output_filename
    )


@app.route('/api/validate', methods=['POST'])
def api_validate():
    """API: Valida datos antes del procesamiento final."""
    data = request.get_json() or {}
    validation_rules = data.get('rules', {})

    filepath = session.get('current_file')
    if not filepath:
        return jsonify({'success': False, 'error': 'No hay archivo cargado'}), 400

    try:
        df = pd.read_excel(filepath)
        errors = []

        # Validaciones básicas
        if len(df) == 0:
            errors.append('El archivo no contiene datos')

        # Validar columnas requeridas según mapeo
        if 'required_columns' in validation_rules:
            for col in validation_rules['required_columns']:
                if col not in df.columns:
                    errors.append(f'Columna requerida faltante: {col}')

        # Validar tipos de datos
        if 'numeric_columns' in validation_rules:
            for col in validation_rules['numeric_columns']:
                if col in df.columns:
                    non_numeric = df[col].apply(lambda x: not isinstance(x, (int, float)) and not pd.isna(x))
                    if non_numeric.any():
                        errors.append(f'Columna {col} contiene valores no numéricos')

        return jsonify({
            'success': len(errors) == 0,
            'valid': len(errors) == 0,
            'errors': errors
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# =============================================================================
# MANEJO DE ERRORES
# =============================================================================

@app.errorhandler(413)
def too_large(e):
    return jsonify({'success': False, 'error': 'Archivo demasiado grande (máx 50MB)'}), 413


@app.errorhandler(500)
def server_error(e):
    return jsonify({'success': False, 'error': 'Error interno del servidor'}), 500


# =============================================================================
# PUNTO DE ENTRADA
# =============================================================================

if __name__ == '__main__':
    print("=" * 60)
    print("  EXPORTADOR INTELIGENTE DE INVENTARIOS")
    print("  Accede a: http://localhost:5000")
    print("=" * 60)
    app.run(debug=True, host='0.0.0.0', port=5000)
